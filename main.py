from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QTextEdit, \
    QFileDialog
import sys
import openpyxl
import urllib3
from funcs.cad import cadFuncionario
from funcs.exc import deleteFunc
from funcs.edit import editFuncionario
from funcs.checar import checar

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

ip = '192.168.1.99'


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()

        # Configuração dos botões
        button1 = QPushButton("Cadastrar", self)
        button1.clicked.connect(self.cadastrar)

        button2 = QPushButton("Editar", self)
        button2.clicked.connect(self.editar)

        button3 = QPushButton("Deletar", self)
        button3.clicked.connect(self.deletar)

        # Configuração da caixa de texto

        self.output_text = QTextEdit(self)
        self.output_text.setReadOnly(True)
        font = self.output_text.font()
        font.setPointSize(9)
        self.output_text.setFont(font)

        # Configuração do layout
        vbox = QVBoxLayout()
        hbox = QHBoxLayout()
        self.setMinimumWidth(500)

        hbox.addWidget(button1)
        hbox.addWidget(button2)
        hbox.addWidget(button3)

        vbox.addLayout(hbox)
        vbox.addWidget(QLabel("Output:"))
        vbox.addWidget(self.output_text)

        self.setLayout(vbox)
        self.setWindowTitle("Controle em massa")
        self.show()

    def selPlanilha(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Selecionar arquivo", "", "Planilha (*.xlsx)")

        if file_path:
            print(f"Caminho do arquivo selecionado: {file_path}")
            return file_path

    def cadastrar(self):
        self.output_text.clear()
        self.output_text.insertPlainText('CADASTRAR🧾\n')
        caminho = self.selPlanilha()
        workbook = openpyxl.load_workbook(caminho)
        sheet = workbook.worksheets[0]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nome = row[0]
            cpf = row[1]
            enable = row[2]
            tag = row[3]
            grupo1 = row[4]
            grupo2 = row[5]
            grupo3 = row[6]
            grupo4 = row[7]
            grupo5 = row[8]
            grupo6 = row[9]
            out = cadFuncionario(ip=ip, nome=nome, cpf=cpf, tag=tag, enable=enable, grupo1=grupo1, grupo2=grupo2,
                                 grupo3=grupo3,
                                 grupo4=grupo4, grupo5=grupo5, grupo6=grupo6)
            while True:
                resp = checar(ip)
                QApplication.processEvents()
                if resp == 'Idle' or resp == 'finished':
                    break

            self.output_text.insertPlainText(f"{out}|{nome} | {tag} | {enable}\n")
            self.output_text.insertPlainText(
                "--------------------------------------------------------------------------------\n")
            self.output_text.verticalScrollBar().setValue(self.output_text.verticalScrollBar().maximum())
            QApplication.processEvents()
        self.output_text.insertPlainText("FINALIZADO✅")

    def editar(self):
        self.output_text.clear()
        self.output_text.insertPlainText("EDITAR📝\n")
        caminho = self.selPlanilha()
        workbook = openpyxl.load_workbook(caminho)
        sheet = workbook.worksheets[0]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nome = row[0]
            cpf = row[1]
            enable = row[2]
            tag = row[3]
            grupo1 = row[4]
            grupo2 = row[5]
            grupo3 = row[6]
            grupo4 = row[7]
            grupo5 = row[8]
            grupo6 = row[9]
            out = editFuncionario(ip=ip, nome=nome, cpf=cpf, tag=tag, enable=enable, grupo1=grupo1,
                                  grupo2=grupo2,
                                  grupo3=grupo3,
                                  grupo4=grupo4, grupo5=grupo5, grupo6=grupo6)
            while True:
                resp = checar(ip)
                QApplication.processEvents()
                if resp == 'Idle' or resp == 'finished':
                    break
            self.output_text.insertPlainText(f"{out}|{nome} \n")
            self.output_text.insertPlainText(
                "--------------------------------------------------------------------------------\n")
            self.output_text.verticalScrollBar().setValue(self.output_text.verticalScrollBar().maximum())
            QApplication.processEvents()

        self.output_text.insertPlainText("FINALIZADO✅")

    def deletar(self):
        self.output_text.clear()
        self.output_text.insertPlainText('DELETAR🛑\n')
        caminho = self.selPlanilha()
        workbook = openpyxl.load_workbook(caminho)
        sheet = workbook.worksheets[0]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nome = row[0]
            tag = row[3]
            out = deleteFunc(ip=ip, nome=nome)
            while True:
                resp = checar(ip)
                QApplication.processEvents()
                if resp == 'Idle' or resp == 'finished':
                    break
            self.output_text.insertPlainText(f"{out}|{nome} | {tag}\n")
            self.output_text.insertPlainText(
                "--------------------------------------------------------------------------------\n")
            self.output_text.verticalScrollBar().setValue(self.output_text.verticalScrollBar().maximum())
            QApplication.processEvents()

        self.output_text.insertPlainText("FINALIZADO✅")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    sys.exit(app.exec_())
