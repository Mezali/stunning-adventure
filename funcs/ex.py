import openpyxl
import time
import urllib3
import funcs.cad as cadastrar
import funcs.exc as excluir
import funcs.edit as editar

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Abre o arquivo da planilha
workbook = openpyxl.load_workbook('C:/Users/autec/PycharmProjects/apiFuncionarios/excel/editFun.xlsx')

# Seleciona a planilha desejada
sheet = workbook.worksheets[0]

resp = input('Cadastrar, Editar ou Deletar?... [C/E/D]\n')

if resp == 'C':
    for row in sheet.iter_rows(min_row=2, values_only=True):
        nome = row[0]
        tag = row[1]
        habilitado = row[2]
        grupo = row[3]
        time.sleep(0.1)
        cadastrar.cadFuncionario(nome=nome, enable=habilitado, tag=tag, grupo1=grupo)

        print(f"{nome} | {tag} | {habilitado}\n")

elif resp == 'E':
    for row in sheet.iter_rows(min_row=2, values_only=True):
        nome = row[0]
        tag = row[1]
        habilitado = row[2]
        grupo = row[3]
        time.sleep(0.5)
        editar.editFuncionario(nome=nome, enable=habilitado, grupo1=grupo, tag=tag)
        print(f"EDITADO(A): {nome}\n")

elif resp == 'D':
    for row in sheet.iter_rows(min_row=2, values_only=True):
        nome = row[0]
        time.sleep(0.5)
        excluir.deleteAllFunc(nome)
        print(f"DELETADO(A): {nome}\n")
else:
    print("DIGITE 'C' PARA CADASTRAR OU 'D' PARA DELETAR😡🤢🤮")
